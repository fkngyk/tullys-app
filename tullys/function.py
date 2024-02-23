import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import datetime
from datetime import datetime
import jpholiday
import calendar
import locale
import os
import math

def input(n, data, year, month, day):
    #日本語に設定
    locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
    
    #原本のファイルを読み込む
    wb1 = xl.load_workbook('/home/fkngyk/fkngyk.pythonanywhere.com/media/files/原本.xlsx')
    ws1 = wb1.worksheets[0]
    ws2 = wb1.worksheets[1]
    
    open = [[0 for i in range(2)] for j in range(n)]
    close = [[0 for i in range(2)] for j in range(n)]
    day_2 = 0
    cnt = [0 for i in range(n)]
    #シフト表の日付などの書き換え
    for i in range(n):
        ws1 = wb1.copy_worksheet(wb1.worksheets[0])
        
        #日付の書き込み
        ws1.cell(1,3).value = str(year) + "年" + str(month) + "月" + str(day) + "日"
        from datetime import date
        the_day = date(year,month,day)
        ws1.title = str(day) + "日"
        
        #条件付き書式の設定
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        dxf1 = DifferentialStyle(fill=blue_fill)
        rule1 = Rule(type="expression", dxf=dxf1)
        rule1.formula = ['R28="o"']
        ws1.conditional_formatting.add("R28:CG37", rule1)
        rule2 = Rule(type="expression", dxf=dxf1)
        rule2.formula = ['R45="o"']
        ws1.conditional_formatting.add("R45:CG68", rule2)
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        dxf2 = DifferentialStyle(fill=yellow_fill)
        rule3 = Rule(type="expression", dxf=dxf2)
        rule3.formula = ['R28="y"']
        ws1.conditional_formatting.add("R28:CG37", rule3)
        rule4 = Rule(type="expression", dxf=dxf2)
        rule4.formula = ['R45="y"']
        ws1.conditional_formatting.add("R45:CG68", rule4)
        
        
        
        #オープン、クローズの時間を設定
        if jpholiday.is_holiday(the_day) == True:
            ws1.cell(1,18).font = xl.styles.fonts.Font(color='FF0000',size=12)
            ws1.sheet_properties.tabColor = 'FF0000'
            
            t = str(ws2.cell(16,3).value)
            o = t.split(":")
            open[day_2][0] = int(o[0])
            open[day_2][1] = int(o[1])
            t = str(ws2.cell(16,5).value)
            c = t.split(":")
            close[day_2][0] = int(c[0])
            close[day_2][1] = int(c[1])
            
            for j in range(18):
                ws1.cell(9,14 + j*4).value = ws2.cell(16,7 + j*4).value
        elif the_day.weekday() == 5:
            t = str(ws2.cell(14,3).value)
            o = t.split(":")
            open[day_2][0] = int(o[0])
            open[day_2][1] = int(o[1])
            t = str(ws2.cell(14,5).value)
            c = t.split(":")
            close[day_2][0] = int(c[0])
            close[day_2][1] = int(c[1])
            ws1.sheet_properties.tabColor = '00CCFF'
            ws1['R1'].font = xl.styles.fonts.Font(color='00CCFF', size=12)
            for j in range(18):
                ws1.cell(9,14 + j*4).value = ws2.cell(14,7 + j*4).value
        elif the_day.weekday() == 6:
            t = str(ws2.cell(16,3).value)
            o = t.split(":")
            open[day_2][0] = int(o[0])
            open[day_2][1] = int(o[1])
            t = str(ws2.cell(16,5).value)
            c = t.split(":")
            close[day_2][0] = int(c[0])
            close[day_2][1] = int(c[1])
            ws1.sheet_properties.tabColor = 'FF0000'
            ws1['R1'].font = xl.styles.fonts.Font(color='FF0000', size=12)
            for j in range(18):
                ws1.cell(9,14 + j*4).value = ws2.cell(16,7 + j*4).value
        else:
            t = str(ws2.cell(4+the_day.weekday()*2,3).value)
            o = t.split(":")
            open[day_2][0] = int(o[0])
            open[day_2][1] = int(o[1])
            t = str(ws2.cell(4+the_day.weekday()*2,5).value)
            c = t.split(":")
            close[day_2][0] = int(c[0])
            close[day_2][1] = int(c[1])
            for j in range(18):
                ws1.cell(9,14 + j*4).value = ws2.cell(4 + the_day.weekday()*2,7 + j*4).value
        
        day += 1
        day_2 += 1
                
    #詳細シートを末尾に移動し、いらないシートを削除
    wb1.move_sheet(ws2,offset=17)
    wb1.remove(wb1.worksheets[0])
    
    for i in range(len(data)):
        for j in range(n):
            start = [0 for k in range(2)]
            finish = [0 for k in range(2)]
            
            #シフトインする日にち、開始時刻、終了時刻を取得
            if data[i][j+1]:
                s = data[i][j+1].split('-')
                date = j
                s2 = math.modf(float(s[0]))
                start[0] = int(s2[1])
                start[1] = int(s2[0]*60)
                if start[0] < open[date][0]:
                    for k in range(2):
                        start[k] = open[date][k]
                elif start[0] == open[date][0] and start[1] < open[date][1]:
                    for k in range(2):
                        start[k] = open[date][k]
                s2 = math.modf(float(s[1]))
                finish[0] = int(s2[1])
                finish[1] = int(s2[0]*60)
                if finish[0] > close[date][0]:
                    for k in range(2):
                        finish[k] = close[date][k]
                elif finish[0] == close[date][0] and finish[1] > close[date][1]:
                    for k in range(2):
                        finish[k] = close[date][k]
        
                ws1 = wb1.worksheets[date+1]
                
                #土日、祝日の場合の営業時間の差分
                x = ((open[date][0]-6)*60 + open[date][1]-30) / 15
                    
                #何マス塗るか計算
                if open[date][0] > start[0]:
                    for k in range(2):
                        start[k] = open[date][k]
                if open[date][0]==start[0] and open[date][1]>start[1]:
                    start[1] = open[date][1]

                if finish[0] > close[date][0]:
                    for k in range(2):
                        finish[k] = close[date][k]
                if close[date][0]==finish[0] and close[date][1]<finish[1]:
                    start[1] = open[date][1]
                    
                t1 = int((start[0]*60+start[1] - open[date][0]*60-open[date][1])/15) + 1
                t2 = int((finish[0]*60+finish[1] - start[0]*60-start[1])/15) + t1 + 1
                    
                #名前の書き込み
                ws1.cell(45+2*cnt[date],1).value = data[i][0].name
                
                #シフト希望時間の書き込み
                for k in range(int(t1)+19,int(t2)+18):
                    ws1.cell(45+2*cnt[date],k+x).value = "o"
                
                cnt[date] += 1
    
    return wb1